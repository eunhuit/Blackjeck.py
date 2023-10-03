import random
from os import system
from time import sleep
from openpyxl import Workbook, load_workbook



# 카드 덱 초기화
suits = ['♠', '♥', '♦', '♣']
ranks = ['2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A']
deck = [(suit, rank) for suit in suits for rank in ranks]

# 점수 계산 함수
def calculate_score(hand):
    score = sum([int(card[1]) if card[1].isdigit() else 10 if card[1] != 'A' else 11 for card in hand])
    if 'A' in [card[1] for card in hand] and score > 21:
        score -= 10
    return score

# 카드 뽑기 함수
def draw_card():
    return deck.pop(random.randint(0, len(deck)-1))

# 게임 시작
def blackjack_game():
    #excel 불러오기
    wb = load_workbook(filename="blackjeck.xlsx")
    ws = wb.active
    cell = ws['A1']
    cell = cell.value
    if cell == None:
        print("초기 자본금 500만원 지원!")
        ws['A1']= "500"
        wb.save('blackjeck.xlsx')
        wb = load_workbook(filename="blackjeck.xlsx")
        ws = wb.active
        cell = ws['A1']
        cell = cell.value
    else:
        pass
    print(f"현재 자본: {cell}만원")
    bat = input("배팅금을 입력해 주세요(단위:만원): ")
    bat = int(bat)
    if int(bat) > int(cell):
        print("자금이 부족합니다.")
        blackjack_game()
    
    # 초기 카드 분배
    player_hand = [draw_card(), draw_card()]
    dealer_hand = [draw_card(), draw_card()]
    
    # 인슈어런스 가능 여부 확인
    #insurance = False
    #if dealer_hand[1][1] == 'A':
    #    insurance = True
    #    print("인슈어런스 가능")
    
    # 게임 진행
    game_over = False
    while not game_over:
        cell = int(cell)
        bat = int(bat)
        player_score = calculate_score(player_hand)
        dealer_score = calculate_score(dealer_hand)
        sleep(2)
        print(f"\n플레이어 카드: {player_hand}, 합계: {player_score}")
        print(f"딜러 카드: {dealer_hand[0]}")
        
        # 플레이어의 선택
        if player_score == 21 and len(player_hand) == 2:
            print("블랙잭! 플레이어 승리!")
            print(f"{bat*2}만원 획득!")
            ws['A1'] = cell + bat*2
            wb.save('blackjeck.xlsx')
            game_over = True
        else:
            choice = input("카드를 더 받으시겠습니까? (Hit/Stand/SPlit/Double): ").lower()
            if choice == 'hit':
                player_hand.append(draw_card())
                player_score = calculate_score(player_hand)
                if player_score > 21:
                    print(f"\n플레이어 카드: {player_hand}, 합계: {player_score}")
                    print("버스트! 딜러 승리!")
                    print(f"{bat}만원을 잃으셨어요")
                    ws['A1'] = cell - bat
                    wb.save('blackjeck.xlsx')
            elif choice == 'h':
                player_hand.append(draw_card())
                player_score = calculate_score(player_hand)
                if player_score > 21:
                    print(f"\n플레이어 카드: {player_hand}, 합계: {player_score}")
                    print("버스트! 딜러 승리!")
                    print(f"{bat}만원을 잃으셨어요")
                    ws['A1'] = cell - bat
                    wb.save('blackjeck.xlsx')
                    game_over = True
            elif choice == 'stand':
                while dealer_score < 17:
                    dealer_hand.append(draw_card())
                    dealer_score = calculate_score(dealer_hand)
                print(f"\n딜러 카드: {dealer_hand}, 합계: {dealer_score}")
                
                if dealer_score > 21:
                    print("딜러 버스트! 플레이어 승리!")
                    print(f"{bat*2}만원 획득!")
                    ws['A1'] = cell + bat*2
                    wb.save('blackjeck.xlsx')
                elif player_score > dealer_score:
                    print("플레이어 승리!")
                    print(f"{bat*2}만원 획득!")
                    ws['A1'] = cell + bat*2
                    wb.save('blackjeck.xlsx')
                elif player_score < dealer_score:
                    print("딜러 승리!")
                    print(f"{bat}만원을 잃으셨어요")
                    ws['A1'] = cell - bat
                    wb.save('blackjeck.xlsx')
                else:
                    print("무승부!")
                    print(f"{bat}만원 푸시")
                    wb.save('blackjeck.xlsx')
                game_over = True
            elif choice == 's':
                while dealer_score < 17:
                    dealer_hand.append(draw_card())
                    dealer_score = calculate_score(dealer_hand)
                print(f"\n딜러 카드: {dealer_hand}, 합계: {dealer_score}")
                
                if dealer_score > 21:
                    print("딜러 버스트! 플레이어 승리!")
                    print(f"{bat*2}만원 획득!")
                    ws['A1'] = cell + bat*2
                    wb.save('blackjeck.xlsx')
                elif player_score > dealer_score:
                    print("플레이어 승리!")
                    print(f"{bat*2}만원 획득!")
                    ws['A1'] = cell + bat*2
                    wb.save('blackjeck.xlsx')
                elif player_score < dealer_score:
                    print("딜러 승리!")
                    print("버스트! 딜러 승리!")
                    print(f"{bat}만원을 잃으셨어요")
                    ws['A1'] = cell - bat
                    wb.save('blackjeck.xlsx')
                else:
                    print("무승부!")
                    print(f"{bat}만원 푸시")
                    wb.save('blackjeck.xlsx')
                game_over = True
            elif choice == 'split':
                # 스플릿 기능 추가
                if player_hand[0][1] == player_hand[1][1]:
                    player_hand_2 = [player_hand.pop()]
                    player_hand.append(draw_card())
                    player_hand_2.append(draw_card())
                    print(f"\n첫 번째 핸드: {player_hand}, 합계: {calculate_score(player_hand)}")
                    print(f"두 번째 핸드: {player_hand_2}, 합계: {calculate_score(player_hand_2)}")
                    print("\n첫 번째 핸드로 게임을 진행합니다.")
                    player_hand = player_hand_2
                else:
                    print("스플릿할 수 없습니다.")
            elif choice == 'sp':
                # 스플릿 기능 추가
                if player_hand[0][1] == player_hand[1][1]:
                    player_hand_2 = [player_hand.pop()]
                    player_hand.append(draw_card())
                    player_hand_2.append(draw_card())
                    print(f"\n첫 번째 핸드: {player_hand}, 합계: {calculate_score(player_hand)}")
                    print(f"두 번째 핸드: {player_hand_2}, 합계: {calculate_score(player_hand_2)}")
                    print("\n첫 번째 핸드로 게임을 진행합니다.")
                    player_hand = player_hand_2
                else:
                    print("스플릿할 수 없습니다.")
            elif choice == 'double':
                # 더블 다운 기능 추가
                if len(player_hand) == 2:
                    player_hand.append(draw_card())
                    player_score = calculate_score(player_hand)
                    print(f"\n플레이어 카드: {player_hand}, 합계: {player_score}")
                    if player_score > 21:
                        print("버스트! 딜러 승리!")
                        print(f"{bat}만원을 잃으셨어요")
                        ws['A1'] = cell - bat
                        wb.save('blackjeck.xlsx')
                        game_over = True
                    else:
                        while dealer_score < 17:
                            dealer_hand.append(draw_card())
                            dealer_score = calculate_score(dealer_hand)
                        print(f"\n딜러 카드: {dealer_hand}, 합계: {dealer_score}")
                        
                        if dealer_score > 21:
                            print("딜러 버스트! 플레이어 승리!")
                            print(f"{bat*2}만원 획득!")
                            ws['A1'] = cell + bat*2
                            wb.save('blackjeck.xlsx')
                        elif player_score > dealer_score:
                            print("플레이어 승리!")
                            print(f"{bat*2}만원 획득!")
                            ws['A1'] = cell + bat*2
                            wb.save('blackjeck.xlsx')
                        elif player_score < dealer_score:
                            print("딜러 승리!")
                            print(f"{bat}만원을 잃으셨어요")
                            ws['A1'] = cell - bat
                            wb.save('blackjeck.xlsx')
                        else:
                            print("무승부!")
                            print(f"{bat}만원 푸시")
                            wb.save('blackjeck.xlsx')
                        game_over = True
                else:
                    print("더블 다운할 수 없습니다.")
            elif choice == 'd':
                # 더블 다운 기능 추가
                if len(player_hand) == 2:
                    player_hand.append(draw_card())
                    player_score = calculate_score(player_hand)
                    print(f"\n플레이어 카드: {player_hand}, 합계: {player_score}")
                    if player_score > 21:
                        print("버스트! 딜러 승리!")
                        print(f"{bat}만원을 잃으셨어요")
                        ws['A1'] = cell - bat
                        wb.save('blackjeck.xlsx')
                        game_over = True
                    else:
                        while dealer_score < 17:
                            dealer_hand.append(draw_card())
                            dealer_score = calculate_score(dealer_hand)
                        print(f"\n딜러 카드: {dealer_hand}, 합계: {dealer_score}")
                        
                        if dealer_score > 21:
                            print("딜러 버스트! 플레이어 승리!")
                            print(f"{bat*2}만원 획득!")
                            ws['A1'] = cell + bat*2
                            wb.save('blackjeck.xlsx')
                        elif player_score > dealer_score:
                            print("플레이어 승리!")
                            print(f"{bat*2}만원 획득!")
                            ws['A1'] = cell + bat*2
                            wb.save('blackjeck.xlsx')
                        elif player_score < dealer_score:
                            print("딜러 승리!")
                            print(f"{bat}만원을 잃으셨어요")
                            ws['A1'] = cell - bat
                            wb.save('blackjeck.xlsx')
                        else:
                            print("무승부!")
                            print(f"{bat}만원 푸시")
                            wb.save('blackjeck.xlsx')
                        game_over = True
                else:
                    print("더블 다운할 수 없습니다.")
            else:
                print("잘못된 입력입니다.")

# 게임 실행
while(True):
    a = input("이어서 진행하시겠습니까? (y/n): ")
    if a == "y":
        system('cls')
        blackjack_game()
    else:
        break

